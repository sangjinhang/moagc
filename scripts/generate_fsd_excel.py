"""
从 COSMIC 功能点 Excel 生成功能结构图 Excel（一级含二级 + 二级含三级）。
"""

import os
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_fsd_excel(input_xlsx: str, output_xlsx: str) -> dict:
    """
    从 COSMIC Excel 生成功能结构图 Excel（一级含二级 + 二级含三级）。

    Args:
        input_xlsx: COSMIC 7列 Excel 文件路径
        output_xlsx: 输出功能结构图 Excel 文件路径

    Returns:
        dict: {
            'l1_count': 一级模块数,
            'l2_count': 二级模块数,
            'sheet_names': [str],  # 所有Sheet名，前l1_count个为一级含二级
        }
    """
    import pandas as pd

    df = pd.read_excel(input_xlsx, sheet_name=0)
    level_cols = ["一级模块", "二级模块", "三级模块"]
    for col in level_cols:
        df[col] = df[col].ffill()

    # 构建层级映射
    l1_to_l2: dict[str, list[str]] = OrderedDict()
    l2_to_l3: dict[str, list[str]] = OrderedDict()

    for _, row in df.iterrows():
        l1, l2, l3 = row["一级模块"], row["二级模块"], row["三级模块"]
        if l1 not in l1_to_l2:
            l1_to_l2[l1] = []
        if l2 not in l1_to_l2[l1]:
            l1_to_l2[l1].append(l2)
        if l2 not in l2_to_l3:
            l2_to_l3[l2] = []
        if l3 not in l2_to_l3[l2]:
            l2_to_l3[l2].append(l3)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_names = []

    # 样式
    title_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    title_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    cell_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    def _write_sheet(wb, sheet_name, parent_name, child_list, per_row=7):
        ws = wb.create_sheet(title=sheet_name[:31])
        sheet_names.append(sheet_name)

        # 标题行
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=per_row + 1)
        title_cell = ws.cell(row=1, column=2, value=f"{parent_name} - 功能模块清单")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_align
        for c in range(2, per_row + 2):
            ws.cell(row=1, column=c).fill = title_fill

        # 表头行（第2行隐藏）
        ws.row_dimensions[2].hidden = True
        ws.cell(row=2, column=1, value=parent_name)

        # A列隐藏
        ws.column_dimensions["A"].hidden = True

        # 写入子模块
        for idx, child in enumerate(child_list):
            r = idx // per_row + 3
            c = idx % per_row + 2
            cell = ws.cell(row=r, column=c, value=child)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = cell_fill

        # 设置列宽
        ws.column_dimensions["B"].width = 28
        for col_letter in [chr(ord("B") + i) for i in range(1, per_row)]:
            ws.column_dimensions[col_letter].width = 28

        # 设置行高
        for r in range(3, (len(child_list) // per_row) + 4):
            ws.row_dimensions[r].height = 36

        ws.row_dimensions[1].height = 40

        # 隐藏网格线
        ws.sheet_view.showGridLines = False

    # 写入一级含二级（排在前面）
    for l1, l2_list in l1_to_l2.items():
        _write_sheet(wb, l1, l1, l2_list)

    # 写入二级含三级（排在后面）
    for l2, l3_list in l2_to_l3.items():
        _write_sheet(wb, l2, l2, l3_list)

    wb.save(output_xlsx)
    return {
        "l1_count": len(l1_to_l2),
        "l2_count": len(l2_to_l3),
        "sheet_names": sheet_names,
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("用法: python generate_fsd_excel.py <输入.xlsx> <输出.xlsx>")
        sys.exit(1)
    result = generate_fsd_excel(sys.argv[1], sys.argv[2])
    print(f"生成完成: 一级模块 {result['l1_count']} 个, 二级模块 {result['l2_count']} 个")


def generate_fsd_excel_from_mapping(
    l1_to_l2: dict, l2_to_l3: dict, output_xlsx: str
) -> dict:
    """
    从层级映射字典生成功能结构图 Excel。

    Args:
        l1_to_l2: 一级模块→二级模块列表映射
        l2_to_l3: 二级模块→三级模块列表映射
        output_xlsx: 输出 Excel 文件路径

    Returns:
        dict: {'l1_count', 'l2_count', 'sheet_names'}
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_names = []

    title_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    title_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    cell_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    per_row = 7

    def _write_sheet(wb, sheet_name, parent_name, child_list):
        ws = wb.create_sheet(title=sheet_name[:31])
        sheet_names.append(sheet_name)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=per_row + 1)
        title_cell = ws.cell(row=1, column=2, value=f"{parent_name} - 功能模块清单")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_align
        for c in range(2, per_row + 2):
            ws.cell(row=1, column=c).fill = title_fill
        ws.row_dimensions[2].hidden = True
        ws.cell(row=2, column=1, value=parent_name)
        ws.column_dimensions["A"].hidden = True
        for idx, child in enumerate(child_list):
            r = idx // per_row + 3
            c = idx % per_row + 2
            cell = ws.cell(row=r, column=c, value=child)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = cell_fill
        ws.column_dimensions["B"].width = 28
        for col_letter in [chr(ord("B") + i) for i in range(1, per_row)]:
            ws.column_dimensions[col_letter].width = 28
        for r in range(3, (len(child_list) // per_row) + 4):
            ws.row_dimensions[r].height = 36
        ws.row_dimensions[1].height = 40
        ws.sheet_view.showGridLines = False

    for l1, l2_list in l1_to_l2.items():
        _write_sheet(wb, l1, l1, l2_list)
    for l2, l3_list in l2_to_l3.items():
        _write_sheet(wb, l2, l2, l3_list)

    wb.save(output_xlsx)
    return {
        "l1_count": len(l1_to_l2),
        "l2_count": len(l2_to_l3),
        "sheet_names": sheet_names,
    }
