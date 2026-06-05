"""
功能结构图 Excel 转 PNG 图片生成器。
每个 Sheet 生成一张 PNG 图片，文件名 = Sheet名.png。
"""

import os
from PIL import Image, ImageDraw, ImageFont
import openpyxl


def convert_fsd_to_images(
    input_xlsx: str,
    output_dir: str,
    per_row: int = 7,
    sheet_filter: list = None,
) -> dict:
    """
    将功能结构图 Excel 转换为 PNG 图片。

    Args:
        input_xlsx: 输入 Excel 文件路径
        output_dir: 输出图片目录路径
        per_row: 每行显示的模块数量，默认 7
        sheet_filter: 仅处理指定的 Sheet 名称列表，为 None 时处理全部

    Returns:
        dict: {'images': [图片路径列表], 'sheet_names': [Sheet名列表]}
    """
    wb = openpyxl.load_workbook(input_xlsx)
    os.makedirs(output_dir, exist_ok=True)

    # 样式参数
    TITLE_BG = (0, 112, 192)
    WHITE = (255, 255, 255)
    BORDER_COLOR = (180, 180, 180)
    TEXT_COLOR = (51, 51, 51)
    CELL_BG = (245, 247, 250)
    CELL_W = 320
    CELL_H = 150
    PADDING = 36
    TITLE_H = 80

    font_title = _load_font("msyhbd.ttc", 32)
    font_cell = _load_font("msyh.ttc", 32)
    font_cell_small = _load_font("msyh.ttc", 24)

    image_paths = []
    sheet_names = []

    # 确定要处理的 Sheet 列表
    target_sheets = sheet_filter if sheet_filter else wb.sheetnames

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # 收集数据行（从第3行开始，B~H列）
        rows_data = []
        for row in ws.iter_rows(
            min_row=3, max_row=ws.max_row,
            min_col=2, max_col=2 + per_row, values_only=True
        ):
            row_vals = [str(v).strip() if v and str(v).strip() else None for v in row]
            if any(row_vals):
                rows_data.append(row_vals)

        if not rows_data:
            continue

        # 计算画布尺寸
        canvas_w = per_row * CELL_W + PADDING * 2 + 10
        canvas_h = PADDING + TITLE_H + PADDING + len(rows_data) * CELL_H + PADDING

        img = Image.new("RGB", (canvas_w, canvas_h), WHITE)
        draw = ImageDraw.Draw(img)

        # 绘制标题栏
        title_text = ws["B1"].value or sheet_name
        title_y = PADDING
        draw.rectangle([PADDING, title_y, PADDING + per_row * CELL_W, title_y + TITLE_H], fill=TITLE_BG)
        _draw_centered_text(draw, title_text, font_title, WHITE, PADDING, title_y, per_row * CELL_W, TITLE_H)

        # 绘制单元格网格
        y_start = title_y + TITLE_H + PADDING
        for r_idx, row_vals in enumerate(rows_data):
            y = y_start + r_idx * CELL_H
            for c_idx in range(per_row):
                x = PADDING + c_idx * CELL_W
                val = row_vals[c_idx] if c_idx < len(row_vals) else None

                if val:
                    draw.rectangle(
                        [x, y, x + CELL_W, y + CELL_H],
                        fill=CELL_BG, outline=BORDER_COLOR
                    )
                    _draw_fitted_text(draw, val, font_cell, font_cell_small, TEXT_COLOR, x, y, CELL_W, CELL_H)
                else:
                    draw.rectangle(
                        [x, y, x + CELL_W, y + CELL_H],
                        fill=WHITE, outline=BORDER_COLOR
                    )

        # 绘制最外层蓝色闭合表框
        tbl_l = PADDING
        tbl_t = PADDING
        tbl_r = PADDING + per_row * CELL_W
        tbl_b = y_start + len(rows_data) * CELL_H
        bw = 3
        draw.line([(tbl_l, tbl_t), (tbl_r, tbl_t)], fill=TITLE_BG, width=bw)
        draw.line([(tbl_l, tbl_b), (tbl_r, tbl_b)], fill=TITLE_BG, width=bw)
        draw.line([(tbl_l, tbl_t), (tbl_l, tbl_b)], fill=TITLE_BG, width=bw)
        draw.line([(tbl_r, tbl_t), (tbl_r, tbl_b)], fill=TITLE_BG, width=bw)

        # 保存
        safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace("*", "").replace("?", "")[:80]
        img_path = os.path.join(output_dir, f"{safe_name}.png")
        img.save(img_path)
        image_paths.append(img_path)
        sheet_names.append(sheet_name)

    return {"images": image_paths, "sheet_names": sheet_names}


def _load_font(font_name: str, size: int):
    try:
        return ImageFont.truetype(font_name, size)
    except (OSError, IOError):
        try:
            return ImageFont.truetype("msyh.ttc", size)
        except (OSError, IOError):
            return ImageFont.load_default()


def _draw_centered_text(draw, text, font, fill, x, y, w, h):
    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    draw.text((x + (w - tw) // 2, y + (h - th) // 2), text, fill=fill, font=font)


def _draw_fitted_text(draw, text, font_normal, font_small, fill, x, y, w, h):
    """绘制文字，优先使用大字体，超宽时折行显示"""
    max_width = w - 16
    bbox = draw.textbbox((0, 0), text, font=font_normal)
    tw = bbox[2] - bbox[0]
    line_height = bbox[3] - bbox[1]

    if tw <= max_width:
        draw.text((x + (w - tw) // 2, y + (h - line_height) // 2), text, fill=fill, font=font_normal)
        return

    lines = _wrap_text(draw, text, font_normal, max_width)
    total_h = line_height * len(lines)

    if total_h <= h - 10:
        _draw_multiline_centered(draw, lines, font_normal, fill, x, y, w, h, line_height)
        return

    bbox2 = draw.textbbox((0, 0), text, font=font_small)
    line_height_s = bbox2[3] - bbox2[1]
    lines_s = _wrap_text(draw, text, font_small, max_width)
    total_h_s = line_height_s * len(lines_s)
    _draw_multiline_centered(draw, lines_s, font_small, fill, x, y, w, h, line_height_s)


def _wrap_text(draw, text, font, max_width):
    """将文字按max_width自动折行"""
    chars = list(text)
    lines = []
    current = ""
    for ch in chars:
        test = current + ch
        bbox = draw.textbbox((0, 0), test, font=font)
        if bbox[2] - bbox[0] > max_width and current:
            lines.append(current)
            current = ch
        else:
            current = test
    if current:
        lines.append(current)
    return lines if lines else [""]


def _draw_multiline_centered(draw, lines, font, fill, x, y, w, h, line_height):
    """多行文字整体居中绘制"""
    total_h = line_height * len(lines)
    start_y = y + (h - total_h) // 2
    for i, line in enumerate(lines):
        bbox = draw.textbbox((0, 0), line, font=font)
        lw = bbox[2] - bbox[0]
        ly = start_y + i * line_height
        draw.text((x + (w - lw) // 2, ly), line, fill=fill, font=font)
